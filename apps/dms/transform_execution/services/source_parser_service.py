"""Parsers de archivo origen según SourceProfile."""

from __future__ import annotations

import csv
import io
import json
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path

from apps.dms.transform_execution.services.capture_boundary_service import (
    apply_row_limit,
    find_line_bounds,
)
from apps.dms.transform_execution.services import source_field_validation_service as field_validation


class ParseError(Exception):
    pass


@dataclass
class ParsedRow:
    line: int
    data: dict


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)
    messages: list[dict] = field(default_factory=list)
    rows_read: int = 0
    start_line: int = 1


def _encoding(source: dict) -> str:
    code = (source.get("encoding_code") or "utf-8").strip().lower()
    if code in {"latin-1", "latin1", "iso-8859-1", "windows-1252", "cp1252"}:
        return "latin-1"
    if code in {"ascii", "utf8", "utf-8", "auto", ""}:
        return "utf-8"
    return code


def _read_text(path: Path, source: dict) -> str:
    raw = path.read_bytes()
    enc = _encoding(source)
    try:
        return raw.decode(enc)
    except UnicodeDecodeError:
        return raw.decode("utf-8", errors="replace")


def _parsing_source(source: dict) -> dict:
    merged = dict(source or {})
    config = source.get("config") or {}
    for key in (
        "delimiter",
        "quote_char",
        "escape_char",
        "has_header",
        "header_row",
        "sheet_name",
        "record_path",
        "record_element",
    ):
        if key in config and key not in merged:
            merged[key] = config[key]
    return merged


def _prepare_text_lines(
    lines: list[str], source: dict
) -> tuple[list[tuple[int, str]], list[dict], list[dict]]:
    """Aplica captura y reglas globales. Devuelve filas, errores y mensajes de captura."""
    rules = source.get("content_rules") or {}
    bounds = find_line_bounds(
        lines,
        source.get("capture_start"),
        source.get("capture_end"),
    )
    captured = list(
        enumerate(lines[bounds.begin : bounds.finish], start=bounds.begin + 1)
    )
    captured = apply_row_limit(captured, source.get("capture_end"))

    comment_prefix = (rules.get("comment_prefix") or "").strip()
    prepared: list[tuple[int, str]] = []
    errors: list[dict] = []

    for line_no, line in captured:
        if rules.get("skip_empty_lines") and not line.strip():
            continue
        if comment_prefix and line.lstrip().startswith(comment_prefix):
            continue

        working = line.strip() if rules.get("trim_lines") else line
        rule_errors = field_validation.line_content_rule_errors(
            working, rules, line_no=line_no
        )
        if rule_errors:
            errors.extend(rule_errors)
            continue
        prepared.append((line_no, working))
    return prepared, errors, list(bounds.messages)


def _parse_fixed_row(line: str, fields: list[dict]) -> dict:
    row = {}
    for field_def in fields:
        name = (field_def.get("name") or "").strip()
        if not name:
            continue
        start = field_def.get("start")
        length = field_def.get("length")
        end = field_def.get("end")
        char = (field_def.get("char") or "").strip()
        try:
            if char and start is None:
                pos = line.find(char)
                if pos < 0:
                    row[name] = ""
                    continue
                s = pos + len(char)
                if length is not None:
                    e = s + int(length)
                elif end is not None:
                    e = int(end)
                else:
                    e = len(line)
                row[name] = line[s:e]
                row[name] = row[name].rstrip()
                continue
            if start is not None and length is not None:
                s = int(start) - 1
                e = s + int(length)
            elif start is not None and end is not None:
                s = int(start) - 1
                e = int(end)
            else:
                row[name] = ""
                continue
            row[name] = line[s:e] if s >= 0 else ""
            row[name] = row[name].rstrip()
        except (TypeError, ValueError, IndexError):
            row[name] = ""
    return row


def _parse_delimited_parts(parts: list[str], fields: list[dict]) -> dict:
    row = {}
    for field_def in fields:
        name = (field_def.get("name") or "").strip()
        if not name:
            continue
        idx = field_def.get("column_index")
        try:
            index = int(idx) if idx is not None else None
        except (TypeError, ValueError):
            index = None
        if index is None or index < 0 or index >= len(parts):
            row[name] = ""
        else:
            row[name] = parts[index]
    return row


def _expected_column_count(fields: list[dict]) -> int | None:
    max_idx = -1
    for field_def in fields or []:
        try:
            idx = int(field_def.get("column_index"))
        except (TypeError, ValueError):
            continue
        max_idx = max(max_idx, idx)
    return max_idx + 1 if max_idx >= 0 else None


def _accept_row(
    fields: list[dict],
    row: dict,
    *,
    line_no: int,
    result: ParseResult,
) -> None:
    result.rows_read += 1
    errors = field_validation.validate_row(fields, row, line_no=line_no)
    if errors:
        result.errors.extend(errors)
        return
    result.rows.append(ParsedRow(line=line_no, data=row))


def _parse_text_rows(path: Path, source: dict, fields: list[dict]) -> ParseResult:
    text = _read_text(path, source)
    lines = text.splitlines()
    prepared, rule_errors, capture_messages = _prepare_text_lines(lines, source)
    result = ParseResult(errors=rule_errors, messages=capture_messages)
    if prepared:
        result.start_line = prepared[0][0]

    file_type = (source.get("file_type_code") or "").strip()
    if file_type == "txt_fixed":
        required_len = field_validation.required_line_length(fields)
        for line_no, line in prepared:
            if not line:
                continue
            if required_len is not None and len(line) < required_len:
                result.rows_read += 1
                result.errors.append(
                    {
                        "line": line_no,
                        "field": "",
                        "code": "LINE_LENGTH_MISMATCH",
                        "message": (
                            f"Línea más corta que la definición posicional "
                            f"({len(line)} < {required_len})."
                        ),
                        "value": line,
                        "actual": len(line),
                        "expected": required_len,
                    }
                )
                continue
            row = _parse_fixed_row(line, fields)
            _accept_row(fields, row, line_no=line_no, result=result)
        return result

    if file_type in {"csv", "txt_delimited"}:
        delimiter = source.get("delimiter") or ","
        if delimiter in {"\\t", "tab"}:
            delimiter = "\t"
        quote = source.get("quote_char") or '"'
        has_header = bool(source.get("has_header"))
        expected_cols = _expected_column_count(fields)
        # Parse línea a línea para conservar número de línea original.
        first_data = True
        for line_no, line in prepared:
            reader = csv.reader(io.StringIO(line), delimiter=delimiter, quotechar=quote)
            try:
                parts = next(reader)
            except StopIteration:
                parts = []
            if has_header and first_data:
                first_data = False
                continue
            first_data = False
            if expected_cols is not None and len(parts) < expected_cols:
                result.rows_read += 1
                result.errors.append(
                    {
                        "line": line_no,
                        "field": "",
                        "code": "DELIMITER_MISMATCH",
                        "message": (
                            f"Columnas inconsistentes ({len(parts)} < {expected_cols})."
                        ),
                        "value": line,
                        "actual": len(parts),
                        "expected": expected_cols,
                    }
                )
                continue
            row = _parse_delimited_parts(parts, fields)
            _accept_row(fields, row, line_no=line_no, result=result)
        return result

    raise ParseError(f"Tipo de origen «{file_type or '—'}» no soportado.")


def _json_path_parts(path: str) -> list[str]:
    text = (path or "").strip()
    if text.startswith("$."):
        text = text[2:]
    elif text.startswith("$"):
        text = text[1:].lstrip(".")
    return [part for part in text.split(".") if part]


def _json_traverse(data, path: str):
    current = data
    for part in _json_path_parts(path):
        key = part.replace("[]", "")
        if not key:
            continue
        if isinstance(current, dict):
            current = current.get(key)
        elif isinstance(current, list):
            if not current:
                return []
            if isinstance(current[0], dict):
                current = [item.get(key) for item in current if isinstance(item, dict)]
            else:
                return []
        else:
            return None
    return current


def _json_field_value(record, json_path: str) -> str:
    current = record
    for part in _json_path_parts(json_path):
        key = part.replace("[]", "")
        if not key:
            continue
        if isinstance(current, dict):
            current = current.get(key)
        else:
            return ""
    if current is None:
        return ""
    if isinstance(current, (dict, list)):
        return json.dumps(current, ensure_ascii=False)
    return str(current)


def _parse_json_rows(path: Path, source: dict, fields: list[dict]) -> ParseResult:
    text = _read_text(path, source)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ParseError(f"JSON inválido: {exc}") from exc

    record_path = (source.get("record_path") or "").strip()
    records = _json_traverse(payload, record_path) if record_path else payload
    if records is None:
        records = []
    if isinstance(records, dict):
        records = [records]
    if not isinstance(records, list):
        raise ParseError("La ruta de registros JSON no devolvió una lista u objeto.")

    capture_end = source.get("capture_end") or {}
    if (capture_end.get("mode") or "").strip() == "max_rows":
        records = apply_row_limit(records, capture_end)

    result = ParseResult(start_line=1)
    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            result.rows_read += 1
            result.errors.append(
                {
                    "line": index,
                    "field": "",
                    "code": "CONTENT_TYPE_MISMATCH",
                    "message": "Registro JSON no es un objeto.",
                    "value": "",
                }
            )
            continue
        row = {}
        for field_def in fields:
            name = (field_def.get("name") or "").strip()
            if not name:
                continue
            json_path = (field_def.get("json_path") or name).strip()
            row[name] = _json_field_value(record, json_path)
        _accept_row(fields, row, line_no=index, result=result)
    return result


def _local_tag(tag: str) -> str:
    return tag.split("}")[-1] if "}" in tag else tag


def _parse_xml_rows(path: Path, source: dict, fields: list[dict]) -> ParseResult:
    record_element = (source.get("record_element") or "").strip()
    if not record_element:
        raise ParseError("Defina record_element para archivos XML.")

    try:
        tree = ET.parse(path)
    except ET.ParseError as exc:
        raise ParseError(f"XML inválido: {exc}") from exc

    root = tree.getroot()
    records = [
        element
        for element in root.iter()
        if _local_tag(element.tag) == record_element
    ]
    if not records and _local_tag(root.tag) == record_element:
        records = [root]

    capture_end = source.get("capture_end") or {}
    if (capture_end.get("mode") or "").strip() == "max_rows":
        records = apply_row_limit(records, capture_end)

    result = ParseResult(start_line=1)
    for index, record in enumerate(records, start=1):
        row = {}
        for field_def in fields:
            name = (field_def.get("name") or "").strip()
            if not name:
                continue
            element_name = (field_def.get("element") or name).strip()
            value = ""
            for child in record:
                if _local_tag(child.tag) == element_name:
                    value = (child.text or "").strip()
                    break
            row[name] = value
        _accept_row(fields, row, line_no=index, result=result)
    return result


def _column_letter_to_index(column: str) -> int | None:
    text = (column or "").strip().upper()
    if not text:
        return None
    if text.isdigit():
        try:
            return max(0, int(text) - 1)
        except ValueError:
            return None
    value = 0
    for char in text:
        if not ("A" <= char <= "Z"):
            return None
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value - 1 if value > 0 else None


def _cell_as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _xlsx_row_bounds(source: dict, max_row: int) -> tuple[int, int, list[dict]]:
    """Filas Excel 1-based [begin, finish] inclusive para datos + mensajes captura."""
    from apps.dms.transform_execution.services.capture_boundary_service import (
        short_file_warning,
    )
    from apps.dms.transform_execution.services.capture_params import (
        capture_line,
        capture_percent,
        normalize_capture,
    )

    start = normalize_capture(source.get("capture_start"))
    end = normalize_capture(source.get("capture_end"))
    start_mode = (start.get("mode") or "first").strip()
    end_mode = (end.get("mode") or "eof").strip()
    messages: list[dict] = []

    begin = 1
    if start_mode == "line_number":
        start_line = capture_line(start) or 1
        begin = max(1, start_line)
        if start_line > max_row:
            messages.append(
                {
                    "level": "warning",
                    "code": "CAPTURE_OUT_OF_RANGE",
                    "expected_line": start_line,
                    "actual_lines": max_row,
                    "text": (
                        f"El archivo tiene {max_row} fila(s); "
                        f"capture_start.line={start_line} queda fuera de rango. "
                        f"No hay filas de captura."
                    ),
                }
            )
    elif start_mode == "after_header_block":
        try:
            begin = max(1, int(start.get("skip_lines") or 0) + 1)
        except (TypeError, ValueError):
            begin = 1
    elif start_mode == "first":
        begin = 1

    header_row = source.get("header_row")
    try:
        header_row_n = int(header_row) if header_row not in (None, "") else None
    except (TypeError, ValueError):
        header_row_n = None
    if header_row_n is not None and begin <= header_row_n:
        begin = header_row_n + 1

    finish = max_row
    if end_mode == "line_number":
        end_line = capture_line(end)
        if end_line is not None:
            if max_row < end_line:
                messages.append(
                    short_file_warning(expected_line=end_line, actual_lines=max_row)
                )
            finish = min(max_row, end_line)
    elif end_mode == "line_or_eof":
        end_line = capture_line(end)
        if end_line is not None:
            finish = min(max_row, end_line)
    elif end_mode == "max_rows":
        try:
            max_rows = int(end.get("max_rows") or 0)
        except (TypeError, ValueError):
            max_rows = 0
        if max_rows > 0:
            finish = min(max_row, begin + max_rows - 1)
    elif end_mode == "percent":
        percent = capture_percent(end) or 100
        finish = min(max_row, max(begin, int(max_row * percent / 100)))

    begin = max(1, min(begin, max_row if max_row else 1))
    finish = max(begin - 1, min(finish, max_row if max_row else 0))
    return begin, finish, messages


def _parse_xlsx_rows(path: Path, source: dict, fields: list[dict]) -> ParseResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ParseError(
            "Falta la dependencia openpyxl para leer archivos Excel."
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"No se pudo abrir el Excel: {exc}") from exc

    try:
        sheet_name = (source.get("sheet_name") or "").strip()
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ParseError(
                    f"No existe la hoja «{sheet_name}». "
                    f"Disponibles: {', '.join(workbook.sheetnames)}."
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        max_row = sheet.max_row or 0
        if max_row < 1:
            return ParseResult(start_line=1)

        begin, finish, capture_messages = _xlsx_row_bounds(source, max_row)
        header_map: dict[str, int] = {}
        header_row = source.get("header_row")
        try:
            header_row_n = int(header_row) if header_row not in (None, "") else None
        except (TypeError, ValueError):
            header_row_n = None
        if header_row_n is not None and 1 <= header_row_n <= max_row:
            header_values = next(
                sheet.iter_rows(
                    min_row=header_row_n, max_row=header_row_n, values_only=True
                ),
                (),
            )
            for col_idx, cell in enumerate(header_values):
                label = _cell_as_text(cell).strip().upper()
                if label:
                    header_map[label] = col_idx

        result = ParseResult(start_line=begin, messages=capture_messages)
        skip_empty = bool((source.get("content_rules") or {}).get("skip_empty_lines", True))
        for row_number, row_cells in enumerate(
            sheet.iter_rows(min_row=begin, max_row=finish, values_only=True),
            start=begin,
        ):
            values = list(row_cells)
            if skip_empty and all(
                value is None or str(value).strip() == "" for value in values
            ):
                continue

            row: dict = {}
            for field_def in fields:
                name = (field_def.get("name") or "").strip()
                if not name:
                    continue
                column = (field_def.get("column") or "").strip()
                index = _column_letter_to_index(column)
                if index is None and column:
                    index = header_map.get(column.upper())
                if index is None or index < 0 or index >= len(values):
                    row[name] = ""
                else:
                    row[name] = _cell_as_text(values[index])
            _accept_row(fields, row, line_no=row_number, result=result)
        return result
    finally:
        workbook.close()


def parse_source_file(path: Path, source: dict, *, limit: int | None = None) -> ParseResult:
    source = _parsing_source(source)
    fields = source.get("fields") or []
    file_type = (source.get("file_type_code") or "").strip()
    if not path.is_file():
        raise ParseError("No se encontró el archivo de entrada.")

    if file_type == "json":
        result = _parse_json_rows(path, source, fields)
    elif file_type == "xml":
        result = _parse_xml_rows(path, source, fields)
    elif file_type == "xlsx":
        result = _parse_xlsx_rows(path, source, fields)
    else:
        result = _parse_text_rows(path, source, fields)

    if limit is not None:
        result.rows = result.rows[: max(0, int(limit))]
    return result

"""Serialización de filas destino según TargetProfile."""

from __future__ import annotations

import csv
import io
import json
import xml.etree.ElementTree as ET


class SerializeError(Exception):
    pass


def _as_text(value, serialization: dict) -> str:
    if value is None or value == "":
        return str(serialization.get("null_representation") or "")
    text = str(value)
    if serialization.get("trim_before_write", True):
        text = text.strip()
    return text


def _apply_field_serialization(text: str, field: dict, global_ser: dict) -> str:
    ser = {**global_ser, **(field.get("serialization") or {})}
    text = _as_text(text, ser)
    max_length = field.get("max_length")
    truncate = ser.get("default_truncate") or ser.get("truncate") or "right"
    if max_length:
        try:
            limit = int(max_length)
        except (TypeError, ValueError):
            limit = None
        if limit and len(text) > limit:
            if truncate == "left":
                text = text[-limit:]
            elif truncate == "error":
                raise SerializeError(
                    f"Campo «{field.get('name')}» supera max_length={limit}."
                )
            else:
                text = text[:limit]
    return text


def _path_parts(path: str) -> list[str]:
    text = (path or "").strip()
    if text.startswith("$."):
        text = text[2:]
    elif text.startswith("$"):
        text = text[1:].lstrip(".")
    return [part for part in text.split(".") if part]


def _set_nested(container: dict, path: str, value) -> None:
    parts = _path_parts(path)
    if not parts:
        return
    current = container
    for part in parts[:-1]:
        next_node = current.get(part)
        if not isinstance(next_node, dict):
            next_node = {}
            current[part] = next_node
        current = next_node
    current[parts[-1]] = value


def _json_value(text: str, field: dict, serialization: dict):
    """Convierte texto serializado a tipo JSON según data_type."""
    data_type = (field.get("data_type") or "string").strip()
    if text == "":
        if data_type in {"integer", "decimal", "float", "boolean"}:
            return None
        return ""
    if data_type == "integer":
        try:
            return int(text)
        except (TypeError, ValueError):
            return text
    if data_type in {"decimal", "float"}:
        try:
            return float(text.replace(",", "."))
        except (TypeError, ValueError):
            return text
    if data_type == "boolean":
        lowered = text.strip().lower()
        if lowered in {"1", "true", "t", "yes", "si", "sí"}:
            return True
        if lowered in {"0", "false", "f", "no"}:
            return False
        return text
    return text


def serialize_rows(rows: list[dict], target: dict) -> bytes:
    file_type = (target.get("file_type_code") or "").strip()
    fields = sorted(
        (target.get("fields") or []),
        key=lambda item: (item.get("order") or 0, item.get("name") or ""),
    )
    serialization = target.get("serialization") or {}
    layout = target.get("layout") or {}

    if file_type == "xlsx":
        return _serialize_xlsx(rows, fields, serialization, layout)
    if file_type == "txt_fixed":
        return _serialize_fixed(rows, fields, serialization, layout)
    if file_type in {"csv", "txt_delimited"}:
        return _serialize_delimited(rows, fields, serialization, layout, file_type)
    if file_type == "json":
        return _serialize_json(rows, fields, serialization, layout)
    if file_type == "xml":
        return _serialize_xml(rows, fields, serialization, layout)
    raise SerializeError(f"Tipo de destino «{file_type or '—'}» no soportado en MVP.")


def _excel_column_index(column) -> int:
    """Convierte letra (A, B, AA) o índice 1-based a índice 1-based."""
    if column in (None, ""):
        return 0
    if isinstance(column, int):
        return column if column >= 1 else 0
    text = str(column).strip().upper()
    if text.isdigit():
        try:
            value = int(text)
        except ValueError:
            return 0
        return value if value >= 1 else 0
    index = 0
    for ch in text:
        if not ("A" <= ch <= "Z"):
            return 0
        index = index * 26 + (ord(ch) - 64)
    return index


def _excel_cell_value(text: str, field: dict, serialization: dict):
    """Valor tipado para celda Excel según excel_type / data_type."""
    excel_type = (field.get("excel_type") or "").strip().lower()
    data_type = (field.get("data_type") or "string").strip().lower()
    kind = excel_type or {
        "integer": "number",
        "decimal": "number",
        "float": "number",
        "boolean": "boolean",
        "date": "date",
        "datetime": "date",
    }.get(data_type, "text")

    if text == "":
        return None if kind in {"number", "boolean", "date"} else ""

    if kind in {"number", "general"} and data_type in {"integer", "decimal", "float"}:
        try:
            if data_type == "integer":
                return int(text)
            return float(text.replace(",", "."))
        except (TypeError, ValueError):
            return text
    if kind == "number":
        try:
            if "." in text or "," in text:
                return float(text.replace(",", "."))
            return int(text)
        except (TypeError, ValueError):
            return text
    if kind == "boolean":
        lowered = text.strip().lower()
        if lowered in {"1", "true", "t", "yes", "si", "sí"}:
            return True
        if lowered in {"0", "false", "f", "no"}:
            return False
        return text
    return text


def _serialize_xlsx(rows, fields, serialization, layout) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SerializeError(
            "Falta la dependencia openpyxl para escribir archivos Excel."
        ) from exc

    sheet_name = (layout.get("sheet_name") or "Hoja1").strip() or "Hoja1"
    include_header = bool(layout.get("include_header", True))
    freeze_header = bool(layout.get("freeze_header_row", True))
    auto_width = bool(layout.get("column_width_auto", True))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]

    # Columna por campo: letter/index explícito, o secuencial por order.
    columns: list[tuple[dict, int]] = []
    used: set[int] = set()
    next_col = 1
    for field in fields:
        col_idx = _excel_column_index(field.get("column"))
        if col_idx < 1:
            while next_col in used:
                next_col += 1
            col_idx = next_col
            next_col += 1
        elif col_idx in used:
            raise SerializeError(
                f"Columna Excel duplicada para «{field.get('name')}»: "
                f"{field.get('column')}."
            )
        used.add(col_idx)
        columns.append((field, col_idx))

    row_no = 1
    if include_header:
        for field, col_idx in columns:
            label = (field.get("label") or field.get("name") or "").strip()
            sheet.cell(row=row_no, column=col_idx, value=label)
        if freeze_header:
            sheet.freeze_panes = "A2"
        row_no += 1

    for row in rows:
        for field, col_idx in columns:
            name = (field.get("name") or "").strip().lower()
            text = _apply_field_serialization(row.get(name, ""), field, serialization)
            value = _excel_cell_value(text, field, serialization)
            sheet.cell(row=row_no, column=col_idx, value=value)
        row_no += 1

    if auto_width and columns:
        for field, col_idx in columns:
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in sheet[letter]:
                raw = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(raw))
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _record_dict(row: dict, fields: list[dict], serialization: dict) -> dict:
    record: dict = {}
    for field in fields:
        name = (field.get("name") or "").strip().lower()
        path = (field.get("path") or name).strip() or name
        if not path:
            continue
        text = _apply_field_serialization(row.get(name, ""), field, serialization)
        value = _json_value(text, field, serialization)
        if value is None and serialization.get("omit_null", False):
            continue
        _set_nested(record, path, value)
    return record


def _serialize_json(rows, fields, serialization, layout) -> bytes:
    root_type = (layout.get("root_type") or "array").strip() or "array"
    records_path = (layout.get("records_path") or "").strip()
    pretty = bool(layout.get("pretty_print", True))
    try:
        indent = int(layout.get("indent") if layout.get("indent") is not None else 2)
    except (TypeError, ValueError):
        indent = 2
    if not pretty:
        indent = None

    records = [_record_dict(row, fields, serialization) for row in rows]

    if root_type == "object":
        if records_path:
            payload: dict = {}
            nested = records[0] if len(records) == 1 else records
            _set_nested(payload, records_path, nested)
        elif len(records) == 1:
            payload = records[0]
        else:
            payload = {"records": records}
    else:
        if records_path:
            payload = {}
            _set_nested(payload, records_path, records)
        else:
            payload = records

    body = json.dumps(payload, ensure_ascii=False, indent=indent).encode("utf-8")
    if indent is not None:
        body = body + b"\n"
    if layout.get("include_bom"):
        body = b"\xef\xbb\xbf" + body
    return body


def _xml_local_name(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return "field"
    text = text.replace("/", ".").split(".")[-1]
    cleaned = "".join(ch if (ch.isalnum() or ch in {"_", "-", "."}) else "_" for ch in text)
    if cleaned and cleaned[0].isdigit():
        cleaned = f"f_{cleaned}"
    return cleaned or "field"


def _serialize_xml(rows, fields, serialization, layout) -> bytes:
    root_name = _xml_local_name(layout.get("root_element") or "records")
    record_name = _xml_local_name(layout.get("record_element") or "record")
    namespace = (layout.get("namespace") or "").strip()
    declaration = bool(layout.get("declaration", True))

    if namespace:
        root = ET.Element(root_name, xmlns=namespace)
    else:
        root = ET.Element(root_name)

    for row in rows:
        record_el = ET.SubElement(root, record_name)
        for field in fields:
            name = (field.get("name") or "").strip().lower()
            path = (field.get("path") or name).strip() or name
            child_name = _xml_local_name(path)
            text = _apply_field_serialization(row.get(name, ""), field, serialization)
            child = ET.SubElement(record_el, child_name)
            child.text = text

    tree = ET.ElementTree(root)
    buffer = io.BytesIO()
    tree.write(
        buffer,
        encoding="utf-8",
        xml_declaration=declaration,
        method="xml",
        short_empty_elements=True,
    )
    return buffer.getvalue()


def _serialize_fixed(rows, fields, serialization, layout) -> bytes:
    record_length = layout.get("record_length")
    trailing = bool(layout.get("trailing_newline", True))
    lines = []
    for row in rows:
        buffer = [" "] * (int(record_length) if record_length else 0)
        built_parts = []
        use_positions = any(field.get("start") is not None for field in fields)
        for field in fields:
            name = (field.get("name") or "").strip().lower()
            text = _apply_field_serialization(row.get(name, ""), field, serialization)
            length = field.get("length") or field.get("max_length")
            align = (field.get("align") or "left").strip()
            pad_char = (field.get("pad_char") or " ")[:1] or " "
            try:
                width = int(length) if length is not None else len(text)
            except (TypeError, ValueError):
                width = len(text)
            if align == "right":
                text = text.rjust(width, pad_char)[:width]
            else:
                text = text.ljust(width, pad_char)[:width]
            if use_positions and field.get("start") is not None:
                start = int(field.get("start")) - 1
                end = start + width
                if not buffer:
                    buffer = [" "] * max(end, int(record_length or end))
                if end > len(buffer):
                    buffer.extend([" "] * (end - len(buffer)))
                buffer[start:end] = list(text)
            else:
                built_parts.append(text)
        line = "".join(buffer) if use_positions else "".join(built_parts)
        if record_length:
            line = line[: int(record_length)].ljust(int(record_length))
        lines.append(line)
    body = ("\n".join(lines) + ("\n" if trailing and lines else "")).encode("utf-8")
    if layout.get("include_bom"):
        body = b"\xef\xbb\xbf" + body
    return body


def _serialize_delimited(rows, fields, serialization, layout, file_type) -> bytes:
    delimiter = layout.get("delimiter") or ("," if file_type == "csv" else ";")
    if delimiter == "\\t":
        delimiter = "\t"
    quote = layout.get("quote_char") or '"'
    include_header = bool(layout.get("include_header", True))
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, quotechar=quote, quoting=csv.QUOTE_MINIMAL)
    names = [(field.get("name") or "").strip().lower() for field in fields]
    labels = [(field.get("label") or field.get("name") or "").strip() for field in fields]
    if include_header:
        writer.writerow(labels)
    for row in rows:
        values = [
            _apply_field_serialization(row.get(name, ""), field, serialization)
            for name, field in zip(names, fields)
        ]
        writer.writerow(values)
    body = output.getvalue().encode("utf-8")
    if layout.get("include_bom"):
        body = b"\xef\xbb\xbf" + body
    return body
